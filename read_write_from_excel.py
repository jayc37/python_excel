#! coding: utf-8
import string
import os
from openpyxl import load_workbook,Workbook
from datetime import datetime,timedelta


#khối lệnh này ghi mảng err_list vào file excel chứa trong folder error_file
def write_error_number_excels(pl):
        ws = Workbook()
        wb = ws.active
        wb['A1'] = 'TITTLE1'
        wb['B1'] = 'TITTLE2'
        wb['C1'] = 'TITTLE3'
        wb['D1'] = 'TITTLE4'
        wb['E1'] = 'TITTLE5'
        for i in range(len(pl)):
            w = 2
            w +=i
            wb['A{}'.format(w)] =  pl[i][0]
            wb['B{}'.format(w)] =  pl[i][1]
            wb['C{}'.format(w)] =  pl[i][2]
            wb['D{}'.format(w)] =  pl[i][3]
        ws.save('error_file\\file_excels_name.xlsx')

#khối lệnh này chuyển dữ liệu từ #file_excel chuyển thành mảng
def work_sheet(wsheet):
    data_sheet = []
    col = [] #column in sheet
    for c in range(wsheet.max_column):
        #got alphabels with max_(len)_column found in worksheet
        col.append(string.ascii_uppercase[c])

    for r in range(2,wsheet.max_row):
        data_row = []
        for c in range(len(col)):
            #got values exactly with "sheet[colum-row]"
            data = wsheet['{}{}'.format(col[c],r)].value
            data_row.append(data)
        data_sheet.append(data_row)
    return data_sheet

def get_data(wbook):
    data = []
    # got values with multiple sheet in workbook
    for sheet in wbook.worksheets:
        wsheet = wbook.active
        data_book = work_sheet(sheet)
        data.extend(data_book)
    return data

# Tìm kiếm các file excel trong folder file_excel_here
for root, dirs, files in os.walk("file_excel_here"):
    for file in files: 
            if file.endswith(".xlsx") and '~' not in file: #tìm ra file excel
                x = re.search(r'pt\d.*',file) #so khớp nameoffile
                file_name = os.path.join(root, file)
                wb = load_workbook(file_name)
                if x:
                    data = get_data(wb)
                    name = re.sub('.xlsx','',x.group()) # cắt chuỗi có kí tự 'pt'+number in nameoffile
                    pushda.getacount(data,name,file) # đăng nhập
                    wb.close()
                else:
                    print("name_of_file must end with 'pt' + 'number' like 'example_pt1'")
